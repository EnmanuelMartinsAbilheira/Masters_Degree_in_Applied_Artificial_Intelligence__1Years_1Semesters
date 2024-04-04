"""
    Course Scheduling System
    Authors:
        - Enmanuel Martins (16430)
        - Pedro Ferreira (17029)
        
"""

from pyomo.environ import *
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ConstraintManager:
    def __init__(self, model, class_names, teacher_subjects, subjects_slots_unavailable,
                 teacher_strict_assign, unavailable_classrooms):
        """ 
            A class responsible for managing and creating constraints for the CourseScheduler class.

            This class provides methods to define and handle constraints to be used in the scheduling process
            performed by the CourseScheduler class.
    
        """
        
        self.model = model
        self.class_names = class_names
        self.teacher_subjects = teacher_subjects
        self.subjects_slots_unavailable = subjects_slots_unavailable
        self.teacher_strict_assign = teacher_strict_assign
        self.unavailable_classrooms = unavailable_classrooms
    
    # Note: We used multiple for class_name in self.class_names: loops.
    # Thus making the code more extensive and less efficient. 
    # Instead we could perform the loop only one time for the multiple constraints.
    # But to make the code easier to read and understand, we chose to implement it with several loops.
      
    def create_constraints(self):
        """ Method to create the constraints for the scheduling problem.
            This method defines and adds constraints to the scheduling problem.
        """
        
        # C1. Distinct for each time slot consrtraint.
        def single_assignment_constraint_rule(model, d, h, course_set):
            """ Ensure that each time slot is assigned to at most one class (for each Course),
                enforcing classroom and subject uniqueness.
                Only one var schedule[d,h,s,c] = 1 to a time slot.
            """
            
            return sum(model.schedule[d, h, s, c] for s in course_set for c in model.Classrooms) <= 1
        
        for class_name in self.class_names:
            setattr(
                self.model,
                f"single_assignment_constraint_{class_name.lower()}",
                Constraint(
                    self.model.Days,
                    self.model.Hours,
                    rule=lambda m, d, h: single_assignment_constraint_rule(m, d, h, getattr(self.model, class_name)),
                ),
            )
        
        
        # C2. Classroom capacity constraint.
        def classroom_capacity_rule(model, d, h, c, course_set):
            """ Ensure that the number of enrolled students of a subject scheduled in a classroom 
                at a given time does not exceed the capacity of that classroom.
            """
            
            return sum(model.schedule[d, h, s, c] * model.StudentsEnrolled[s] for s in course_set) \
                <= model.ClassroomCapacities[c]


        for class_name in self.class_names:
            setattr(
                self.model,
                f"classroom_capacity_constraint_{class_name.lower()}",
                Constraint(
                    self.model.Days,
                    self.model.Hours,
                    self.model.Classrooms,
                    rule=lambda m, d, h, c: classroom_capacity_rule(m, d, h, c, getattr(self.model, class_name)),
                ),
            )
        
            
        # C3. Update subject lectured variable constraint.
        def subject_lectured_rule(model, d, s, subj_lectured_var):
            """ Rule to ensure that the subject_lectured[d, s] variable takes the value 1
                if the subject is scheduled for that day.
            """
            return subj_lectured_var[d, s] >= sum(model.schedule[d, h, s, c] for h in model.Hours for c in model.Classrooms)
                
                
        for class_name in self.class_names:
            setattr(
                self.model,
                f"subject_lectured_constraint_{class_name.lower()}",
                Constraint(
                    self.model.Days,
                    getattr(self.model, class_name),
                    rule=lambda m, d, s: subject_lectured_rule(m, d, s,
                                                               getattr(self.model, f"subject_lectured_{class_name.lower()}")),
                )
            )
        
        
        # C4. Update any_classes_on_day[d] variable constraint.
        def any_classes_on_day_constraint_rule(model, d, any_classes_var, course_set):
            """ Rule to ensure that any_classes_on_day[d] variable takes 1 if at least one class
                is scheduled on day d.
                If any classes in day d, any_classes_on_day[d] will be set to 1; Otherwise, it will be 0.
            """
            
            return any_classes_var[d] >= sum(model.schedule[d, h, s, c] for h in model.Hours for s in course_set \
                for c in model.Classrooms) / len(model.Hours)


        for class_name in self.class_names:
            setattr(
                self.model,
                f"any_classes_on_day_constraint_{class_name.lower()}",
                Constraint(
                    self.model.Days,
                    rule=lambda m, d: any_classes_on_day_constraint_rule(m, d, getattr(self.model, \
                        f"any_classes_on_day_{class_name.lower()}"), getattr(self.model, class_name)),
                ),
            )
        
        
        # C5. Constraint for max/min days of classes (for each subject).
        for class_name in self.class_names:
            constraint_list = ConstraintList()
            setattr(self.model, f"subject_days_range_constraint_{class_name.lower()}", constraint_list)
            
            for subject in getattr(self.model, class_name):
                constraint_list.add(
                    sum(getattr(self.model, f"subject_lectured_{class_name.lower()}")[d, subject] for d in self.model.Days)
                    <= self.model.MaxDaysPerSubject[subject]
                )
                
                constraint_list.add(
                    sum(getattr(self.model, f"subject_lectured_{class_name.lower()}")[d, subject] for d in self.model.Days)
                    >= self.model.MinDaysPerSubject[subject]
                )
         
        
        # C6. Update cumulative blocks var constraint
        def update_cumulative_blocks_rule(model, s, cumulative_var):
            """ Rule to update the cumulative_blocks (for each subject) variable based on the schedule. """
            
            return cumulative_var[s] == sum(model.schedule[d, h, s, c] for d in model.Days for h in model.Hours \
                for c in model.Classrooms)
          
            
        for class_name in self.class_names:
            cumulative_var = getattr(self.model, f"cumulative_blocks_{class_name.lower()}")
            
            setattr(self.model, f"update_cumulative_blocks_constraint_{class_name.lower()}", 
                    Constraint(getattr(self.model, class_name), 
                            rule=lambda m, s: update_cumulative_blocks_rule(m, s, cumulative_var))
                )
            
        
        # C7. Max/Min blocks of classes per week for each subject. 
        def min_cumulative_blocks_constraint_rule(model, s, cumulative_var):
            """ Rule for the lower bound on cumulative_blocks. """
            return cumulative_var[s] >= model.MinBlocksPerWeek[s]
         
         
        def max_cumulative_blocks_constraint_rule(model, s, cumulative_var):
            """ Rule for the lower bound on cumulative_blocks. """
            
            return cumulative_var[s] <= model.MaxBlocksPerWeek[s]

        def update_cumulative_vs_subj_lectured_rule(model, s, cumulative_var, subject_lect_var):
            """Rule to update cumulative_blocks regarding subject_lectured variable. """
             
            return cumulative_var[s] == sum(subject_lect_var[d, s] for d in model.Days)
        
    
        for class_name in self.class_names:
            cumulative_var = getattr(self.model, f"cumulative_blocks_{class_name.lower()}")
            subject_lectured_var = getattr(self.model, f"subject_lectured_{class_name.lower()}")
            
            # Min cumulative blocks constraint
            setattr(self.model, f"min_cumulative_blocks_constraint_{class_name.lower()}", 
                    Constraint(getattr(self.model, class_name), 
                            rule=lambda m, s: min_cumulative_blocks_constraint_rule(m, s, cumulative_var))
                )
            
            # Max cumulative blocks constraint
            setattr(self.model, f"max_cumulative_blocks_constraint_{class_name.lower()}", 
                    Constraint(getattr(self.model, class_name), 
                            rule=lambda m, s: max_cumulative_blocks_constraint_rule(m, s, cumulative_var))
                )

            # Update cumulative blocks vs subject_lectured constraint
            setattr(self.model, f"cumulative_blocks_subj_lectured_constraint_{class_name.lower()}", 
                    Constraint(getattr(self.model, class_name), 
                            rule=lambda m, s: update_cumulative_vs_subj_lectured_rule(m, s, cumulative_var, 
                                                                                      subject_lectured_var))
                )

        # C8. Gap_indicator variable constraint.
        def gap_constraint_rule(model, d, h, next_h, gap_var, course_set):
            """ Rule to update gap_indicator variable.
                Indicates gaps in the schedule within the same day.
            """
            
            return gap_var[d, h, next_h] <= sum(model.schedule[d, h, s, c] - model.schedule[d, next_h, s, c] \
                for s in course_set for c in model.Classrooms)
            

        for class_name in self.class_names:
            gap_indicator = getattr(self.model, f"gap_indicator_{class_name.lower()}")
            
            # Gap constraint
            setattr(self.model, f"gap_constraint_{class_name.lower()}", 
                    Constraint(self.model.Days, self.model.ConsecutiveHours, 
                            rule=lambda m, d, h, next_h: gap_constraint_rule(m, d, h, next_h, gap_indicator, getattr(self.model, class_name)))
                )
           
            
        # C9. Teacher unavailability constraint.
        def teacher_unavailability_rule(model, d, h, s, c):
            """ Rule to set the value of teacher_unavailable based on the dict containing
                the teacher unavailabilities.
            """
            
            if s in self.subjects_slots_unavailable and (d, h) in self.subjects_slots_unavailable[s]:
                return model.schedule[d, h, s, c] <= 0
            else:
                return Constraint.Skip

        self.model.teacher_unavailability_rule_constraint = Constraint(
            self.model.Days, self.model.Hours, self.model.Subjects, self.model.Classrooms, rule=teacher_unavailability_rule)
        
        
        # C10. Constraint to set the value of schedule for strict teacher assignments
        self.model.strictassign_constraint = ConstraintList()
        for _, assignments in self.teacher_strict_assign.items():
            for assignment in assignments:
                day, hour, subject = assignment
                self.model.strictassign_constraint.add(expr=sum(self.model.schedule[day, hour, subject, c] for \
                    c in self.model.Classrooms) == 1)
                
                
        # C11. Classrooms unavailability constraint.
        def unavailable_classrooms_constraint_rule(model, d, h, c):
            """ Rule for unavailable classrooms in a certain (day, hour). """
            
            if c in self.unavailable_classrooms:
                for unavailable_tuple in self.unavailable_classrooms[c]:
                    if (d, h) == unavailable_tuple:
                        return sum(model.schedule[d, h, s, c] for s in model.Subjects) == 0
            return Constraint.Skip

        self.model.unavailable_classrooms_constraint = Constraint(self.model.Days, self.model.Hours, self.model.Classrooms, \
            rule=unavailable_classrooms_constraint_rule)
        
        
        # C12. Update adjacent_days variable constraint.
        def adjacent_days_constraint_rule(_model, d, d_next, adjacent_days_var, any_classes_on_day_var):
            """ Rule to ensure adjacent_days[d, d_next] is 1 if both days have at least one class
                scheduled otherwise gets 0. Then try to maximize that situation on the objective.
            """ 
            
            # Note:
            # _model is just a placeholder variable that indicates the presence of the model instance within the constraint rule.
            # It's not directly used within the rule but serves as a convention to indicate that the rule is defined within
            # the context of a Pyomo model.
            return adjacent_days_var[d, d_next] <= any_classes_on_day_var[d] + any_classes_on_day_var[d_next] - 1 
          

        for class_name in self.class_names:
            adjacent_days_var = getattr(self.model, f"adjacent_days_{class_name.lower()}")
            any_classes_on_day_var = getattr(self.model, f"any_classes_on_day_{class_name.lower()}")
            
            # Adjacent days constraint
            setattr(self.model, f"adjacent_days_constraint_{class_name.lower()}", 
                    Constraint(self.model.ConsecutiveDays, 
                            rule=lambda m, d, d_next: adjacent_days_constraint_rule(m, d, d_next, adjacent_days_var, any_classes_on_day_var))
                ) 
        
        
        # C13. Min slots/blocks per day constraint.
        def min_blocks_per_day_with_classes_constraint_rule(model, d, course_set, any_classes_on_day_var):
            """ Rule to ensure at least n blocks of classes (for each course) per day when the day has classes."""
            
            min_hours_per_day_param = getattr(model, f"{course_set}_min_blocks_per_day")
            
            return sum(model.schedule[d, h, s, c] for h in model.Hours for s in course_set for c in model.Classrooms) >= \
                min_hours_per_day_param * any_classes_on_day_var[d]
                

        for class_name in self.class_names:
            any_classes_on_day_var = getattr(self.model, f"any_classes_on_day_{class_name.lower()}")
            
            # Min blocks per day with classes constraint
            setattr(self.model, f"min_blocks_per_day_with_classes_constraint_{class_name.lower()}", 
                    Constraint(self.model.Days, rule=lambda m, d: \
                        min_blocks_per_day_with_classes_constraint_rule(m, d, getattr(m, class_name), any_classes_on_day_var))
                )
            
            
        # C14. Max subject slots/blocks per day constraint.
        def max_subject_blocks_per_day_with_classes_constraint_rule(model, d, s):
            """ Rule to ensure at most MaxSubjectBlocksPerDay blocks per day for each subject. """
            
            return sum(model.schedule[d, h, s, c] for h in model.Hours for c in model.Classrooms) \
                <= model.MaxSubjectBlocksPerDay
        
        
        for class_name in self.class_names:
           
            # Max subject blocks per day with classes constraint
            setattr(self.model, f"max_subject_blocks_per_day_with_classes_constraint_{class_name.lower()}", 
                    Constraint(self.model.Days, getattr(self.model, class_name), rule=lambda m, d, s: \
                        max_subject_blocks_per_day_with_classes_constraint_rule(m, d, s))
                )
            
            
        # C15. Min blocks (2-hours) of classes per period (Morning/Afternoon) when the day has classes.
        def any_classes_day_period_constraint_rule(model, p, d, any_classes_day_period, course_set):
            """ Rule to update any_classes_day_period[p, d] variable. """
            
            hours_set = model.period_sets_dict.get(p, set())
            return any_classes_day_period[p, d] >= sum(model.schedule[d, h, s, c] for h in hours_set\
                for s in course_set for c in model.Classrooms) / len(hours_set)

        for class_name in self.class_names:
            any_classes_day_period_var = getattr(self.model, f"any_classes_day_period_{class_name.lower()}")
            
            # Update any_classes_day_period var constraint
            setattr(self.model, f"any_classes_day_period_constraint_{class_name.lower()}", 
                    Constraint(self.model.Periods, self.model.Days, rule=lambda m, p, d: \
                        any_classes_day_period_constraint_rule(m, p, d, any_classes_day_period_var, getattr(m, class_name)))
                )
        
        
        # C16. Ensure at least min_blocks_per_period block(2h) of classes per period when the day and period has classes
        def min_blocks_per_period_with_classes_constraint_rule(model, d, p, any_classes_day_period, course_set):
            """ Rule to ensure at least min_blocks_per_period (2h blocks) of classes per period 
                when the day and period has classes.
            """
            
            hours_set = model.period_sets_dict.get(p, set())
            min_classes_per_period_param = getattr(model, f"{course_set}_min_blocks_per_period")
            
            return sum(model.schedule[d, h, s, c] for h in hours_set for s in course_set for c in model.Classrooms) >= \
                min_classes_per_period_param * any_classes_day_period[p, d]

        for class_name in self.class_names:
            any_classes_day_period_var = getattr(self.model, f"any_classes_day_period_{class_name.lower()}")
            
            # Min blocks per period constraint
            setattr(self.model, f"min_blocks_per_period_with_classes_constraint_{class_name.lower()}", 
                    Constraint(self.model.Days, self.model.Periods, rule=lambda m, d, p: \
                        min_blocks_per_period_with_classes_constraint_rule(m, d, p, any_classes_day_period_var, getattr(m, class_name)))
                )
        
        
        # C17. Max blocks per day with classes constraint
        def max_blocks_per_day_with_classes_constraint_rule(model, d, any_classes_on_day_var, course_set):
            """ Rule to ensure at most n blocks of classes per day when the day has classes. """
            
            max_hours_per_day_param = getattr(model, f"{course_set}_max_blocks_per_day")
            return sum(model.schedule[d, h, s, c] for h in model.Hours for s in course_set for c in model.Classrooms) <= \
                max_hours_per_day_param * any_classes_on_day_var[d]

        for class_name in self.class_names:
            any_classes_on_day_var = getattr(self.model, f"any_classes_on_day_{class_name.lower()}")
            
            setattr(self.model, f"max_blocks_per_day_with_classes_constraint_{class_name.lower()}", 
                    Constraint(self.model.Days, rule=lambda m, d: \
                        max_blocks_per_day_with_classes_constraint_rule(m, d, any_classes_on_day_var, getattr(m, class_name)))
                )
        
        
        # C18. Max days of classes (all subjects of each Course) constraint.
        def max_days_of_classes_rule(model, any_classes_on_day_var):
            """ Rule to ensure at most n days of classes (all subjects of each Course). """
            
            max_day_of_classes_param = getattr(model, f"{class_name.lower()}_max_days_of_classes")
            return sum(any_classes_on_day_var[d] for d in model.Days) <= max_day_of_classes_param
        
        for class_name in self.class_names:
            any_classes_on_day_var = getattr(self.model, f"any_classes_on_day_{class_name.lower()}")
            
            # Max days of classes constraint
            setattr(self.model, f"max_days_of_classes_constraint_{class_name.lower()}", 
                    Constraint(rule=lambda m: max_days_of_classes_rule(m, any_classes_on_day_var))
                )
        
        
        # C19. Min days of classes (all subjects of each Course) constraint.
        def min_days_of_classes_rule(model, any_classes_on_day_var):
            """ Rule to ensure at least n days of classes (all subjects of each Course). """
            
            min_day_of_classes_param = getattr(model, f"{class_name.lower()}_min_days_of_classes")
            return sum(any_classes_on_day_var[d] for d in model.Days) >= min_day_of_classes_param
        
        for class_name in self.class_names:
            any_classes_on_day_var = getattr(self.model, f"any_classes_on_day_{class_name.lower()}")
            
            # Min days of classes constraint
            setattr(self.model, f"min_days_of_classes_constraint_{class_name.lower()}", 
                    Constraint(rule=lambda m: min_days_of_classes_rule(m, any_classes_on_day_var))
                )
        
        
        # C20. Max blocks/slots per week constraint.
        def max_blocks_per_week_constraint_rule(model, course_set):
            """ Rule to ensure at most n blocks of classes per week (total classes on a week for each Course). """
            
            max_hours_per_week_param = getattr(model, f"{course_set}_max_blocks_per_week")
            total_hours_per_week = sum(model.schedule[d, h, s, c] for d in model.Days for h in model.Hours for s in course_set \
                for c in model.Classrooms)
            
            return total_hours_per_week <= max_hours_per_week_param

        for class_name in self.class_names:         
            # Max blocks per week constraint
            setattr(self.model, f"max_blocks_per_week_constraint_{class_name.lower()}", 
                    Constraint(rule=lambda m : max_blocks_per_week_constraint_rule(m, getattr(m, class_name)))
                )
                     
             
        # C21. Max blocks per teacher in a day constraint.
        def max_blocks_per_teacher_constraint_rule(model, t, d):
            """ Rule to limit the maximum number of blocks per teacher per day (across all Courses)"""
            subjects_taught = self.teacher_subjects.get(t, ())
            
            if isinstance(subjects_taught, str):  # Single subject lectured
                return sum(model.schedule[d, h, subjects_taught, c] for h in model.Hours for c in model.Classrooms)\
                    <= model.MaxBlocksPerTeacher
            elif isinstance(subjects_taught, tuple):  # Multiple subjects lectured
                return sum(model.schedule[d, h, s, c] for s in subjects_taught for h in model.Hours for c in model.Classrooms)\
                    <= model.MaxBlocksPerTeacher

        self.model.max_blocks_per_teacher_constraint = Constraint(self.model.Teachers, self.model.Days, \
            rule=max_blocks_per_teacher_constraint_rule)
        
        
        # C22. Classroom scheduled uniqueness between Courses constraint.
        # One classroom for the classes in a day d, hour h, classroom c (across the n courses / between all courses)
        def distinct_classroom_between_courses(model, d, h, c):
            """ Rule to assure one classroom for the classes in a day d, hour h 
                (across the n courses / between all courses).
            """
            
            # Count the number of classes scheduled for each course in self.class_names
            classes_in_classes = [sum(model.schedule[d, h, s, c] for s in getattr(model, class_name.lower()))\
                for class_name in self.class_names]

            # Ensure that the sum of all is less than or equal to 1 (at most one class between courses)
            return sum(classes_in_classes) <= 1

        self.model.distinct_classroom_between_courses_constraint = Constraint(self.model.Days, self.model.Hours, \
            self.model.Classrooms, rule=distinct_classroom_between_courses)
       
       
        # C23. Teacher assignment uniqueness (between courses) constraint.
        def one_class_at_time_per_teacher(model, d, h, t):
            """ Rule to assure that a subject lectured by same teacher is assigned in different (d, h) between the n courses. """
            
            subjects_taught = self.teacher_subjects.get(t, ())
            
            if isinstance(subjects_taught, str):  # Single subject lectured
                return sum(model.schedule[d, h, subjects_taught, c] for c in model.Classrooms)\
                    <= 1
            elif isinstance(subjects_taught, tuple):  # Multiple subjects lectured
                return sum(model.schedule[d, h, s, c] for s in subjects_taught for c in model.Classrooms)\
                    <= 1
        
        self.model.one_class_at_time_per_teacher_constraint = Constraint(self.model.Days, self.model.Hours, \
            self.model.Teachers, rule=one_class_at_time_per_teacher)
            
        
class WorksheetManager:
    """ 
    A class for managing schedules dataframes and processing the information to save it to Excel files.

    This class provides functionality to handle schedules dataframes and perform operations such as 
    generating Excel files from the data.
    
    """
    
    def __init__(self, classical_schedule_df, all_info_df, filename='', flask_src = True) -> None:
        """ Constructor method for initializing a WorksheetManager instance. """
        
        self.wb = Workbook()
        self.ws = self.wb.active
        
        # Format the first DataFrame starting from cell B2
        self.classical_scheduling_table(classical_schedule_df, start_row=2, start_column=2)
        self.all_info_schedule_table(all_info_df, start_row=2, start_column=8)
        self.set_column_width()
        # Save the workbook to an Excel file
        if flask_src:
            folder_path = os.path.join(os.getcwd(), 'Schedules')
        else:
             folder_path = os.path.join(os.getcwd(), 'Schedules(Main_HardCoded)')
             
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
            
        self.wb.save(filename=os.path.join(folder_path, filename))
        

    def set_column_width(self):
        """
            Adjusts the width of columns in the worksheet based on the maximum length of content in each column.
            This method iterates over each column in the worksheet, calculates the maximum length of content 
            in that column, and adjusts the column width accordingly to ensure all content is fully visible.
            
        """
        for col in self.ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            column = get_column_letter(col[0].column)  # Convert column index to letter
            self.ws.column_dimensions[column].width = adjusted_width 
            
    def common_formatting(self, df, start_row, start_column, data_cell=True, is_index=False):
        """ Apply common formatting to the cells of the worksheet based on the provided DataFrame.
            Common formatting between the 2 distinct df types (Schedule_all_info and Classical_schedule).
            
        """
        
        # For data cells
        if data_cell:
            for r_idx, row in enumerate(df.itertuples(), start=start_row):
                for c_idx, value in enumerate(row[1:], start=start_column):
                    cell = self.ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    thin_border = Border(left=Side(style='hair'), 
                                    right=Side(style='hair'), 
                                    top=Side(style='hair'), 
                                    bottom=Side(style='hair'))
                    cell.border = thin_border
                    fill = PatternFill(start_color='f5f5f5', end_color='f5f5f5', fill_type='solid')
                    cell.fill = fill
        # For headers
        else:
            # Header from index?
            if is_index:
                for r_idx, hour_label in enumerate(df.index, start=start_row):
                    cell = self.ws.cell(row=r_idx, column=start_column - 1, value=hour_label)
                    cell.font = Font(bold=True, color='faf8f7')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    fill = PatternFill(start_color='140601', end_color='140601', fill_type='solid')  
                    cell.fill = fill
                     
            # Header from column?
            else:
                for c_idx, day_label in enumerate(df.columns, start=start_column):
                    cell = self.ws.cell(row=start_row - 1, column=c_idx, value=day_label)
                    cell.font = Font(bold=True, color='faf8f7')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    fill = PatternFill(start_color='140601', end_color='140601', fill_type='solid')
                    cell.fill = fill      
        
        
    def classical_scheduling_table(self, my_df, start_row=1, start_column=1):
        """ Method to format the df/table (Classical_schedule)."""
        
        # Write and format data cells
        self.common_formatting(my_df, start_row, start_column, data_cell=True)
        
        # Write and format hour headers to the table
        self.common_formatting(my_df, start_row, start_column, data_cell=False, is_index=True)
         
        # Write and format day headers to the table
        self.common_formatting(my_df, start_row, start_column, data_cell=False, is_index=False)
        
        # Add additional styling for the "----------" cells and set to 'Lunch'
        for r_idx in range(start_row, self.ws.max_row + 1):
            for c_idx in range(start_column, self.ws.max_column + 1):
                cell_value = self.ws.cell(row=r_idx, column=c_idx).value
                if cell_value == '----------':
                    self.ws.cell(row=r_idx, column=c_idx).value = 'Lunch'
                    self.ws.cell(row=r_idx, column=c_idx).fill = PatternFill(
                        start_color='e5f018', end_color='e5f018', fill_type='solid'
                    )  
                    self.ws.cell(row=r_idx, column=c_idx).font = Font(color='808080')
    
 
    def all_info_schedule_table(self, my_df, start_row=2, start_column=8):
        """ Method to format the df/table (Schedule_all_info)."""
        
        # Write and format data cells
        self.common_formatting(my_df, start_row, start_column, data_cell=True)
        
        # Write and format headers to the table
        self.common_formatting(my_df, start_row, start_column, data_cell=False, is_index=False)

    
    
class CourseScheduler:
    """ 
    A class for solving Course scheduling problems using mathematical optimization.

    This class provides methods to formulate and solve scheduling problems for courses,
    using mathematical optimization techniques.
    
    """
    def __init__(self, classes_dict, days, hours, teacher_subject, teachers, unavailable_classrooms,
                 classroom_capacities, students_enrolled, time_ranges, weights, src_flask=True):
        
        """ Constructor method for initializing a CourseScheduler instance. """
        
        # Subjects for each class/course
        # Will be overridden later in initialize_sets() (transforming attributes values into Pyomo sets)
        for class_name, subj_list in classes_dict.items():
            setattr(self, f'{class_name.lower()}', subj_list)
        
        self.class_names = [key.lower() for key in classes_dict.keys()]
        self.days = days
        self.hours = hours
        self.subjects = []
        
        # All subjects (all courses/classes)
        for subjects_entry in teacher_subject.values():
            if isinstance(subjects_entry, tuple):
                self.subjects.extend(subjects_entry)
            else:
                self.subjects.append(subjects_entry)
    
        self.classrooms = list(classroom_capacities.keys())
        self.teachers = list(teachers.keys())
        self.teacher_subject = teacher_subject

        # To check if the call comes from flask_app or not:
        self.src_flask = src_flask
        
        # Unpack the time_ranges list
        # See the descriptions in the main file (main_hard_coded.py)
        (self.min_days_classes_subject, self.max_days_classes_subject, self.min_blocks_per_week_subject,
         self.max_blocks_per_week_subject, self.max_blocks_per_day_teacher, self.general_course_data) = time_ranges
        
        # Max number of blocks each subject has per day (1 subject appears at most one time per day)
        # The problem was formulated for this value to be 1
        self.max_subject_blocks_per_day = 1
        
        self.unavailable_classrooms = unavailable_classrooms
        self.classroom_capacities = classroom_capacities
        self.students_enrolled = students_enrolled
        
        # Unpack the weights list
        # See the descriptions in the main file (main_hard_coded.py)
        (self.penalty_weight_dispersion, self.gap_penalty_weight, 
        self.gap_between_days) = weights
        
        self.model = ConcreteModel()
        self.initialize_sets()
        self.process_teacher_data(teachers)
        self.initialize_parameters()
        self.initialize_variables()
        self.initialize_objective(self.class_names)
        # Initialize the constraints
        self.constraint_manager = ConstraintManager(self.model, self.class_names, self.teacher_subject, 
                                                    self.subjects_slots_unavailable,
                                                    self.teacher_strict_assign,
                                                    self.unavailable_classrooms).create_constraints()
        
    
    def process_teacher_data(self, teachers): # Note: teachers (dict) from the __init__ entry not self.teachers: list!!!
        """ Method to process the dicts associated with the teachers data/params. """
        
        self.teacher_unavailability = {}
        self.teacher_preferences = {}
        self.teacher_strict_assign = {}
        self.subjects_slots_unavailable = {}

        for teacher, details in teachers.items():
            # Handle teacher unavailability
            unavailability = details.get('unavailability', [])
            self.teacher_unavailability[teacher] = unavailability

            # Handle teacher preferences
            preferences = details.get('preferences', [])
            self.teacher_preferences[teacher] = [
                (day, time, subject, weight) for day, time, subject, weight in preferences
            ]

            # Handle teacher strict assignments
            strict_assign = details.get('strict_assign', [])
            self.teacher_strict_assign[teacher] = strict_assign

            # Map teacher unavailability to unavailability of time slots for the
            # subjects lectured by that teacher
            subjects = self.teacher_subject[teacher]
            if isinstance(subjects, tuple):
                for subject in subjects:
                    if subject not in self.subjects_slots_unavailable:
                        self.subjects_slots_unavailable[subject] = []
                    self.subjects_slots_unavailable[subject].extend(unavailability)
            else:
                if subjects not in self.subjects_slots_unavailable:
                    self.subjects_slots_unavailable[subjects] = []
                self.subjects_slots_unavailable[subjects].extend(unavailability)

        
    def initialize_sets(self):
        """ Initializes the sets required for the course scheduling problem. """
        
        self.model.Days = Set(initialize=self.days)
        self.model.Hours = Set(initialize=self.hours)
        self.model.Subjects = Set(initialize=self.subjects)

        # Set eg: self.model.lesi -> Set[list of lesi subjects] (Pyomo Set)
        for class_name in self.class_names:
            setattr(
                self.model,
                f"{class_name.lower()}",
                Set(initialize=getattr(self, class_name.lower()))
            )
            
        self.model.Classrooms = Set(initialize=self.classrooms)
        self.model.Teachers = Set(initialize=self.teachers)
        self.model.ClassroomPairs = Set(within=self.model.Classrooms * self.model.Classrooms, initialize=lambda model: ((c1, c2) \
            for c1 in model.Classrooms for c2 in model.Classrooms if c1 != c2))
        
        hours = list(self.model.Hours)
        self.model.ConsecutiveHours = Set(within=self.model.Hours * self.model.Hours, initialize=lambda model: ((h1, h2) \
            for h1 in hours for h2 in hours if hours.index(h1) + 1 == hours.index(h2)))
            
        split_hour = 13
        morning_hours = [hour for hour in hours if int(hour.split('-')[0]) < split_hour]
        afternoon_hours = [hour for hour in hours if int(hour.split('-')[0]) >= split_hour]
  
        self.model.Morning = Set(initialize=morning_hours)
        self.model.Afternoon = Set(initialize=afternoon_hours)
        self.model.Periods = Set(initialize=['Morning', 'Afternoon'])
        self.model.period_sets_dict = {
            'Morning': self.model.Morning,
            'Afternoon': self.model.Afternoon,
        }
        
        days = list(self.model.Days)
        self.model.ConsecutiveDays = Set(within=self.model.Days * self.model.Days, initialize=lambda model: ((d1, d2) \
            for d1 in days for d2 in days if days.index(d1) + 1 == days.index(d2)))
    
    
    def initialize_parameters(self):
        """ Initializes the parameters required for the course scheduling problem. """
        
        # Classroom capacities
        self.model.ClassroomCapacities = Param(self.model.Classrooms, initialize=self.classroom_capacities)

        # Number of students enrolled
        self.model.StudentsEnrolled = Param(self.model.Subjects, initialize=self.students_enrolled)

        # Maximum days for classes for each subject (Eg: Math has at most n classes per day)
        self.model.MaxDaysPerSubject = Param(self.model.Subjects, initialize=self.max_days_classes_subject)
        # Minimum days for classes for each subject (Eg: Math has at least n classes per day)
        self.model.MinDaysPerSubject = Param(self.model.Subjects, initialize=self.min_days_classes_subject)

        # Maximum blocks per subject per week
        self.model.MaxBlocksPerWeek = Param(self.model.Subjects, initialize=self.max_blocks_per_week_subject)
        # Minimum blocks per subject per week
        self.model.MinBlocksPerWeek = Param(self.model.Subjects, initialize=self.min_blocks_per_week_subject)

        # Max blocks (2h) of a teacher per day
        self.model.MaxBlocksPerTeacher = Param(initialize=self.max_blocks_per_day_teacher)
        
        # Max subject blocks (2 hour) per day
        self.model.MaxSubjectBlocksPerDay = Param(initialize = self.max_subject_blocks_per_day)
        
        # Initialize general parameters
        # General (all subjects) time related parameters for each course
        # See the descriptions in the main file (main_hard_coded.py)
        for course, params in self.general_course_data.items():
            course_lower = course.lower()
            for param_name, param_value in params.items():
                setattr(self.model, f"{course_lower}_{param_name}", Param(initialize=param_value))

        
        # Particular preferences of the teachers
        self.model.Preferences = Param(self.model.Days, self.model.Hours, self.model.Subjects, initialize=1.0, mutable=True)

        # Update preferences Param
        # _ represents the key of the dict -> teacher (not used in the loop)
        for _, preferences_list in self.teacher_preferences.items():
            for day, hour, subject, weight in preferences_list:
                # Update the Preferences parameter with the weight at the specified indices
                self.model.Preferences[day, hour, subject] = weight
    
    
    def initialize_variables(self):
        """ Initializes the variables required for the course scheduling problem. """
        
        # Decision variable (main var)
        self.model.schedule = Var(self.model.Days, self.model.Hours, self.model.Subjects, self.model.Classrooms, domain=Binary)

        # Aux variables
        # Variables specific to each class/course
        for class_name in self.class_names:
            # Variable to check if subject was scheduled on that day, its a flag like var. Takes the value 1 
            # if a subject is scheduled for that day.
            setattr(
                self.model,
                f"subject_lectured_{class_name.lower()}",
                Var(self.model.Days, getattr(self.model, class_name), domain=Binary)
            )
            # Cumulative blocks variable (each subject)
            # Variable to count the number of total (for the week) slots/blocks assigned for each subject.
            setattr(
                self.model,
                f"cumulative_blocks_{class_name.lower()}",
                Var(getattr(self.model, class_name), domain=NonNegativeReals, initialize=0.0)
            )
            # Variable to check if the day has classes scheduled (that day has classes?).
            setattr(
                self.model,
                f"any_classes_on_day_{class_name.lower()}",
                Var(self.model.Days, domain=Binary)
            )
            # Indicator variable for identifying gaps between classes within a day.
            setattr(
                self.model,
                f"gap_indicator_{class_name.lower()}",
                Var(self.model.Days, self.model.ConsecutiveHours, domain=Binary)
            )
            # Variable for identifying gaps between classes on each pair of consecutive days (adjacent_days).
            # Then on the objective, we try to maximize the number of consecutive days with classes.
            setattr(
                self.model,
                f"adjacent_days_{class_name.lower()}",
                Var(self.model.ConsecutiveDays, domain=Binary)
            )
            # Variable to check if subject was scheduled on that period (Morning/Afternoon).
            # For each day.
            setattr(
                self.model,
                f"any_classes_day_period_{class_name.lower()}",
                Var(self.model.Periods, self.model.Days, domain=Binary)
            )
        
    
    def objectives_per_class(self, class_name):
        """ Initializes the objectives per class/course required for the course scheduling problem. """
        
        any_classes_on_day = getattr(self.model, f'any_classes_on_day_{class_name}')
        gap_indicator = getattr(self.model, f'gap_indicator_{class_name}')
        adjacent_days = getattr(self.model, f'adjacent_days_{class_name}')
        
        return (
                self.penalty_weight_dispersion * sum(
                any_classes_on_day[d] for d in self.model.Days)
                    
                + self.gap_penalty_weight * sum(
                    gap_indicator[d, (h1, h2)]
                    for d in self.model.Days
                    for (h1, h2) in self.model.ConsecutiveHours)
                
                + self.gap_between_days * -sum(
                    adjacent_days[(d1, d2)]
                    for (d1, d2) in self.model.ConsecutiveDays)
            
        )  
        
        
    def initialize_objective(self, class_names_list):
        """ Initializes the global objective method required for the course scheduling problem. """
        
        self.model.obj = Objective(
            expr=-sum(
                self.model.schedule[d, h, s, c] * self.model.Preferences[d, h, s]
                for d in self.model.Days
                for h in self.model.Hours
                for s in self.model.Subjects
                for c in self.model.Classrooms
            )
            + sum(self.objectives_per_class(class_name) for class_name in class_names_list),
            sense=minimize
        )
   
          
    def solve_schedule(self, mysolver='gurobi'):
        """ Method to solve the problem. """ 
           
        opt = SolverFactory(mysolver)
        results = opt.solve(self.model)
        status_condition = 0
        # Check solver results
        if (results.solver.status == SolverStatus.ok
            and results.solver.termination_condition == TerminationCondition.optimal):
            status_condition = 1
            print("Solver found optimal solution.")
            # Print some info (Test/Debug)
            self.print_some_info()
            self.print_and_save_schedules()
            
        elif results.solver.termination_condition == TerminationCondition.infeasible:
            print("Solver found the problem to be infeasible.")
        else:
            print("Solver failed to find an optimal solution.")


        return results, status_condition
        
        
    def print_and_save_schedules(self):
        """ Method to print and save the results in a schedule like structure. """
        
        def map_subject_to_teacher(subject):
            """ Method to to map subjects to teachers. """
            
            for teacher, subjects_taught in self.teacher_subject.items():
                if isinstance(subjects_taught, str) and subjects_taught == subject:
                    return teacher
                elif isinstance(subjects_taught, tuple) and subject in subjects_taught:
                    return teacher
            return None
        
       
        def manage_schedules(class_name, class_set):
            """ Method to create the 2 distinct dataframes (for each class/course). """
            
            schedule_data = [
                (d, h, s, c) for d in self.model.Days for h in self.model.Hours for s in class_set for c in self.model.Classrooms
                if self.model.schedule[d, h, s, c].value > 0.5
            ]

            df_schedule = pd.DataFrame(schedule_data, columns=['Day', 'Hour', 'Subject', 'Classroom'])
            
            # Add teacher to the original dataframe
            df_schedule['Teacher'] = df_schedule['Subject'].map(map_subject_to_teacher)
            df_schedule['Teacher'] = df_schedule.apply(lambda row: row['Teacher'] if isinstance(row['Teacher'], str) else \
                ', '.join(row['Teacher']), axis=1)
            df_schedule = df_schedule[['Day', 'Hour', 'Subject', 'Teacher', 'Classroom']]
            
            # Extract _ from subjects with that character
            df_schedule['Subject'] = df_schedule['Subject'].str.split('_').str[0]
            
            # Create a list to store the new rows
            new_rows = []

            # Iterate through each row and duplicate based on the hour range
            for _, row in df_schedule.iterrows():
                start_hour = row['Hour'].split('-')[0]
                end_hour = row['Hour'].split('-')[1]
                for hour in range(int(start_hour), int(end_hour)):
                    new_row = row.copy()
                    new_row['New_hour'] = f'{hour}-{hour + 1}'
                    new_rows.append(new_row)

            # Create a new DataFrame with the duplicated rows
            df_split_hours = pd.DataFrame(new_rows)

            # Drop the original 'Hour' column and unnecessary columns
            df_split_hours = df_split_hours.drop(['Hour'], axis=1)

            # Reorder columns to match the original order
            df_split_hours = df_split_hours[['Day', 'New_hour', 'Subject', 'Classroom']]

            # Sort the DataFrame by 'Day', 'New_hour', and 'Subject'
            df_split_hours = df_split_hours.sort_values(by=['Day', 'New_hour', 'Subject']).reset_index(drop=True)
                        

            # Pivot table to create classical-like schedule
            df_classical_schedule = pd.pivot_table(df_split_hours, values='Subject', index='New_hour',
                                                   columns='Day', aggfunc='first')

            print(f'{class_name}---------------------------------')
            print(df_schedule)
            print('----------------------------------')

            column_order = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']
            df_classical_schedule = df_classical_schedule.reindex(columns=column_order, fill_value='')

            hour_order = ['9-10', '10-11', '11-12', '12-13', '13-14', '14-15', '15-16', '16-17', '17-18']
            df_classical_schedule = df_classical_schedule.reindex(hour_order)
            df_classical_schedule = df_classical_schedule.fillna('')

            df_classical_schedule.loc['13-14'] = '----------'
            print(f'{class_name}................................')
            print(df_classical_schedule)
            print('----------------------------------')
            
            return df_schedule, df_classical_schedule
        
        
        schedules_dict = {}
        for class_name in self.class_names:
            class_set = getattr(self.model, class_name.lower())

            # Create the dataframes using the manage_schedules function
            df_schedule, df_classical_schedule = manage_schedules(class_name, class_set)
            schedules_dict[class_name] = df_classical_schedule
            
            # Create WorksheetManager instance to save to excel files
            WorksheetManager(df_classical_schedule, df_schedule, \
                filename=f'Schedule_{class_name.capitalize()}.xlsx', flask_src = self.src_flask)
            
        
    def print_some_info(self):
        """ Method to print some info. Test/Debug purposes. """
        
        # Print the objective value
        print('----------------------')
        print(f"Objective Value: {self.model.obj()}")
        print('----------------------')
     
        for class_name in self.class_names:
            # Print cumulative blocks info for each course
            cumulative_blocks_var = getattr(self.model, f"cumulative_blocks_{class_name.lower()}")
            print(f"Cumulative Blocks for {class_name}:")
            for subject in getattr(self, f'{class_name.lower()}'):
                print(f"Subject: {subject}, Value: {value(cumulative_blocks_var[subject])}")
                
            # Print subject_lectured var for each course
            var_name = f"subject_lectured_{class_name.lower()}"
            var = getattr(self.model, var_name)
            print(f"Values of {var_name}:")
            for d in self.model.Days:
                for subj in getattr(self.model, class_name.lower()):
                    print(f"Day: {d}, Subject: {subj}, Value: {value(var[d, subj])}")
             
            """ # Print adjacent_days for each course
            adjacent_days = getattr(self.model, f"adjacent_days_{class_name.lower()}")
            print(f"Ajacent Days {class_name}:")
            for (d1, d2) in self.model.ConsecutiveDays:
                print(f"Adjacent Days: {(d1, d2)}, Value: {value(adjacent_days[(d1, d2)])}") """
                
            """ # Print any_classes_on_day_ for each course
            any_classes = getattr(self.model, f"any_classes_on_day_{class_name.lower()}")
            print(f"Any classes on day for: {class_name}:")
            for d in self.model.Days:
                print(f"Day: {d}, Value: {value(any_classes[d])}")
            
            var_any_period = getattr(self.model, f"any_classes_day_period_{class_name.lower()}")
            for period in self.model.Periods:
                for day in self.model.Days:
                    print(f"Period: {period}, Day: {day}, Value: {value(var_any_period[period, day])}")    
            
            var_gap_indicator = getattr(self.model, f"gap_indicator_{class_name.lower()}")
            for day in self.model.Days:
                for hour1, hour2 in self.model.ConsecutiveHours:
                    print(f"Day: {day}, Hours: {hour1}-{hour2}, Value: {value(var_gap_indicator[day, hour1, hour2])}")  """           